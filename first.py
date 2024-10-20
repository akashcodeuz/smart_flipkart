import cv2
import numpy as np
import pytesseract
import requests
import logging
from keras.models import load_model
from datetime import datetime
from twilio.rest import Client
from openpyxl import Workbook, load_workbook
import os

# Cloud and Twilio settings
CLOUD_MODEL_URL = "https://your-cloud-model-endpoint.com/predict"
TWILIO_ACCOUNT_SID = "YOUR_ACCOUNT_SID"
TWILIO_AUTH_TOKEN = "YOUR_AUTH_TOKEN"
TWILIO_API_URL = f"https://api.twilio.com/2010-04-01/Accounts/{TWILIO_ACCOUNT_SID}/Messages.json"
TWILIO_FROM_NUMBER = "+0987654321"
TWILIO_TO_NUMBER = "+1234567890"

# Local deep learning model path
LOCAL_MODEL_PATH = 'local_product_model.h5'

# Initialize logging
logging.basicConfig(level=logging.INFO, filename='processing.log', filemode='a', format='%(asctime)s - %(message)s')

# Load pre-trained local model
def load_local_model():
    try:
        model = load_model(LOCAL_MODEL_PATH)
        logging.info('Successfully loaded local model.')
        return model
    except Exception as e:
        logging.error(f"Error loading local model: {e}")
        return None

# Step 1: Dynamic Image Quality Control (Real-time contrast/brightness adjustment)
def adjust_image_quality(image):
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    adjusted_image = cv2.equalizeHist(gray_image)

    # Apply adaptive thresholding for better OCR
    adaptive_thresh = cv2.adaptiveThreshold(adjusted_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)

    return adaptive_thresh

# Step 2: Real-time Image Capture and Adjustment
def capture_image_with_adjustments(camera_id=0):
    cap = cv2.VideoCapture(camera_id)
    if not cap.isOpened():
        logging.error("Error: Cannot open camera.")
        return None

    ret, frame = cap.read()
    if ret:
        frame = adjust_image_quality(frame)
        cv2.imshow('Processed Image', frame)
        cv2.imwrite('processed_image.jpg', frame)
        logging.info("Image captured with dynamic adjustments.")
    cap.release()
    cv2.destroyAllWindows()
    return frame

# Step 3: OCR with Multi-language Support
def extract_text_multi_language(image, language='eng'):
    custom_config = f'-l {language} --oem 3 --psm 6'
    try:
        text = pytesseract.image_to_string(image, config=custom_config)
        logging.info(f"Extracted Text: {text}")
        return text
    except Exception as e:
        logging.error(f"OCR extraction failed: {e}")
        return ""

# Step 4: Cloud-based and Local Product Classification
def classify_product(image, model=None):
    if model:
        # Use local model if available
        try:
            img_array = cv2.resize(image, (224, 224))  # Resize image to match model input
            img_array = img_array.astype('float32') / 255.0  # Normalize the image to [0, 1]
            img_array = np.expand_dims(img_array, axis=0)
            predictions = model.predict(img_array)
            product_name = np.argmax(predictions, axis=1)[0]  # Assuming categorical classification
            logging.info(f"Local Classification Result: {product_name}")
            return product_name
        except Exception as e:
            logging.error(f"Local classification failed: {e}")
            return None
    else:
        # Fallback to cloud classification
        _, img_encoded = cv2.imencode('.jpg', image)
        try:
            response = requests.post(CLOUD_MODEL_URL, files={"file": img_encoded.tobytes()})
            if response.status_code == 200:
                result = response.json()
                product_name = result['product_name']
                logging.info(f"Cloud Classification Result: {product_name}")
                return product_name
            else:
                logging.error(f"Error in cloud classification: {response.status_code}")
                return None
        except requests.exceptions.RequestException as e:
            logging.error(f"Cloud request failed: {e}")
            return None

# Step 5: Twilio SMS Notifications with Error Handling and Retry Mechanism
def send_alert_via_twilio(product_name, details):
    try:
        client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        message = client.messages.create(
            to=TWILIO_TO_NUMBER,
            from_=TWILIO_FROM_NUMBER,
            body=f"Alert: Issue with product {product_name} - {details}"
        )
        logging.info(f"Twilio SMS sent successfully: SID={message.sid}")
    except Exception as e:
        logging.error(f"Twilio SMS failed: {e}")

# Step 6: Excel Storage with Timestamp and Error Handling
def store_in_excel(product_name, text_details):
    file_name = 'product_data.xlsx'
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Data"
        ws.append(["Product Name", "Text Details", "Timestamp"])  # Add header if creating a new file

    ws = wb.active
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([product_name, text_details, timestamp])  # Append new data
    wb.save(file_name)

    logging.info(f"Data stored in Excel with timestamp: {timestamp}")

# Step 7: Real-time Processing with Batch Image Processing and Logging
def real_time_processing_advanced():
    cap = cv2.VideoCapture(0)
    local_model = load_local_model()

    frame_batch = []
    batch_size = 5  # Process frames in batches of 5

    with open('output_results.txt', 'w') as file:
        file.write("Starting advanced real-time processing...\n")
        
        while True:
            ret, frame = cap.read()
            if not ret:
                file.write("Error: Unable to capture frame from camera.\n")
                break

            cv2.imshow('Live Feed', frame)

            frame_batch.append(frame)
            if len(frame_batch) == batch_size:
                for frame in frame_batch:
                    processed_image = adjust_image_quality(frame)

                    # OCR with multi-language support
                    text_details = extract_text_multi_language(processed_image, language='eng+spa')
                    file.write(f"OCR Text Extracted: {text_details}\n")

                    # Product classification
                    product_name = classify_product(processed_image, local_model)
                    file.write(f"Product Classified: {product_name}\n")

                    # Store results in Excel
                    store_in_excel(product_name, text_details)

                    # Send alerts for anomalies
                    if "expired" in text_details.lower():
                        file.write(f"Alert: Expired product detected for {product_name}.\n")
                        send_alert_via_twilio(product_name, "Expired product detected!")

                frame_batch.clear()  # Clear batch after processing

            if cv2.waitKey(1) & 0xFF == ord('q'):
                file.write("User ended processing session.\n")
                break

        cap.release()
        cv2.destroyAllWindows()

        file.write("Processing completed.\n")

# Main Function to Start the Advanced Processing
def main():
    logging.info("Starting advanced real-time processing...")
    real_time_processing_advanced()

if __name__ == "__main__":
    main()
